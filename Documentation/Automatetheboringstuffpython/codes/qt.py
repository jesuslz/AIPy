import openpyxl as pyxl 
wb = pyxl.load_workbook('/Users/mirnazertuche/Desktop/example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')

print(sheet.cell(row=1, column=2).value)
a = sheet.max_row
b = sheet.max_column

for i in range(1, b + 1):
    for j in range(1, a + 1):
        print(j, sheet.cell(row = j, column = i).value)
        #print("\n")
sheet.insert_rows(3, amount = 1)
'''
print(sheet['A1'].value)
print(sheet['B1'].value)

print(sheet)
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)
'''
