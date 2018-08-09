import openpyxl

try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
get_column_letter(sheet.max_column) #(sheet.get_highest_column())
column_index_from_string('A')
column_index_from_string('AA')